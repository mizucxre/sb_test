import logging
import os
import uuid
from fastapi import UploadFile, File
from PIL import Image
from fastapi import FastAPI, Depends, HTTPException, Request, Form, Query, Response
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from typing import List, Optional
from datetime import datetime, timedelta
import json
import io

from app.database import db
from app.services.order_service import OrderService, ParticipantService
from app.services.user_service import AddressService, SubscriptionService
from app.services.admin_service import AdminService
from app.services.admin_chat_service import AdminChatService
from app.models import Order, AdminUserCreate, AdminUserUpdate, AdminChatMessageCreate
from app.config import STATUSES
from app.utils.security import verify_password, create_access_token, verify_token, generate_avatar_url
from app.utils.session import get_current_admin  # Убираем require_super_admin

logger = logging.getLogger(__name__)

app = FastAPI(title="SEABLUU Admin", docs_url=None, redoc_url=None)

# Определяем базовые пути для статики и шаблонов
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(BASE_DIR, "static")
TEMPLATES_DIR = os.path.join(BASE_DIR, "templates")

# Mount static files and templates
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
templates = Jinja2Templates(directory=TEMPLATES_DIR)

def serialize_model(model):
    """Сериализация Pydantic модели в словарь с обработкой разных версий Pydantic"""
    try:
        # Пробуем Pydantic v2
        return model.model_dump()
    except AttributeError:
        try:
            # Пробуем Pydantic v1
            return model.dict()
        except AttributeError:
            # Если не Pydantic модель, используем __dict__
            return model.__dict__

# Вспомогательная функция для проверки супер-админа
def check_super_admin(current_admin: dict):
    """Проверка что пользователь супер-админ"""
    if current_admin.get("role") != "super_admin":
        raise HTTPException(status_code=403, detail="Super admin access required")
    return current_admin

# Страница входа
@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    return templates.TemplateResponse("login.html", {
        "request": request
    })

@app.post("/login")
async def login(request: Request, response: Response):
    """Аутентификация администратора"""
    try:
        form_data = await request.form()
        username = form_data.get("username")
        password = form_data.get("password")
        
        if not username or not password:
            raise HTTPException(400, "Необходимо указать имя пользователя и пароль")
        
        # Проверяем учетные данные
        admin_user = await AdminService.authenticate_user(username, password)
        if not admin_user:
            raise HTTPException(401, "Неверное имя пользователя или пароль")
        
        # Создаем токен
        access_token = create_access_token(
            data={"sub": admin_user.username, "user_id": admin_user.id, "role": admin_user.role}
        )
        
        # Устанавливаем cookie
        response = RedirectResponse(url="/admin/", status_code=302)
        response.set_cookie(
            key="admin_token",
            value=access_token,
            httponly=True,
            max_age=60 * 60 * 24 * 7,  # 7 дней
            secure=False,  # Для продакшена установите True
            samesite="lax"
        )
        
        # Обновляем время последнего входа
        await AdminService.update_last_login(admin_user.id)
        
        return response
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Login error: {e}")
        raise HTTPException(500, "Внутренняя ошибка сервера")

@app.get("/logout")
async def logout(response: Response):
    """Выход из системы"""
    response = RedirectResponse(url="/admin/login", status_code=302)
    response.delete_cookie("admin_token")
    return response

# Защищенные страницы
@app.get("/", response_class=HTMLResponse)
async def admin_dashboard(request: Request, current_admin: dict = Depends(get_current_admin)):
    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "current_admin": current_admin,
        "current_page": "dashboard"
    })

@app.get("/orders", response_class=HTMLResponse)
async def orders_page(request: Request, current_admin: dict = Depends(get_current_admin)):
    return templates.TemplateResponse("orders.html", {
        "request": request,
        "current_admin": current_admin,
        "current_page": "orders",
        "statuses": STATUSES
    })

@app.get("/orders/new", response_class=HTMLResponse)
async def new_order_page(request: Request, current_admin: dict = Depends(get_current_admin)):
    return templates.TemplateResponse("order_form.html", {
        "request": request,
        "current_admin": current_admin,
        "statuses": STATUSES,
        "current_page": "orders"
    })

@app.get("/orders/{order_id}/edit", response_class=HTMLResponse)
async def edit_order_page(request: Request, order_id: str, current_admin: dict = Depends(get_current_admin)):
    order = await OrderService.get_order(order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    return templates.TemplateResponse("order_form.html", {
        "request": request,
        "current_admin": current_admin,
        "statuses": STATUSES,
        "current_page": "orders",
        "order": order
    })

@app.get("/participants", response_class=HTMLResponse)
async def participants_page(request: Request, current_admin: dict = Depends(get_current_admin)):
    return templates.TemplateResponse("participants.html", {
        "request": request,
        "current_admin": current_admin,
        "current_page": "participants"
    })

@app.get("/reports", response_class=HTMLResponse)
async def reports_page(request: Request, current_admin: dict = Depends(get_current_admin)):
    return templates.TemplateResponse("reports.html", {
        "request": request,
        "current_admin": current_admin,
        "current_page": "reports"
    })

@app.get("/broadcast", response_class=HTMLResponse)
async def broadcast_page(request: Request, current_admin: dict = Depends(get_current_admin)):
    return templates.TemplateResponse("broadcast.html", {
        "request": request,
        "current_admin": current_admin,
        "current_page": "broadcast",
        "statuses": STATUSES
    })

@app.get("/settings", response_class=HTMLResponse)
async def settings_page(request: Request, current_admin: dict = Depends(get_current_admin)):
    return templates.TemplateResponse("settings.html", {
        "request": request,
        "current_admin": current_admin,
        "current_page": "settings",
        "statuses": STATUSES
    })

# Новые страницы для управления администраторами
@app.get("/admin-users", response_class=HTMLResponse)
async def admin_users_page(request: Request, current_admin: dict = Depends(get_current_admin)):
    # Проверяем права супер-админа вручную
    check_super_admin(current_admin)
    
    return templates.TemplateResponse("admin_users.html", {
        "request": request,
        "current_admin": current_admin,
        "current_page": "admin_users"
    })

@app.get("/admin-users/new", response_class=HTMLResponse)
async def new_admin_user_page(request: Request, current_admin: dict = Depends(get_current_admin)):
    # Проверяем права супер-админа вручную
    check_super_admin(current_admin)
    
    return templates.TemplateResponse("admin_user_form.html", {
        "request": request,
        "current_admin": current_admin,
        "current_page": "admin_users"
    })

@app.get("/admin-users/{user_id}/edit", response_class=HTMLResponse)
async def edit_admin_user_page(request: Request, user_id: int, current_admin: dict = Depends(get_current_admin)):
    # Проверяем права супер-админа вручную
    check_super_admin(current_admin)
    
    user = await AdminService.get_user_by_id(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Admin user not found")
    
    return templates.TemplateResponse("admin_user_form.html", {
        "request": request,
        "current_admin": current_admin,
        "current_page": "admin_users",
        "user": user
    })

@app.get("/admin-chat", response_class=HTMLResponse)
async def admin_chat_page(request: Request, current_admin: dict = Depends(get_current_admin)):
    return templates.TemplateResponse("admin_chat.html", {
        "request": request,
        "current_admin": current_admin,
        "current_page": "admin_chat"
    })

@app.get("/profile", response_class=HTMLResponse)
async def profile_page(request: Request, current_admin: dict = Depends(get_current_admin)):
    return templates.TemplateResponse("profile.html", {
        "request": request,
        "current_admin": current_admin,
        "current_page": "profile"
    })

# API endpoints для администраторов
@app.get("/api/admin/users")
async def get_admin_users(current_admin: dict = Depends(get_current_admin)):
    """Получение списка администраторов"""
    try:
        check_super_admin(current_admin)
        users = await AdminService.get_all_users()
        return {"users": users}
    except Exception as e:
        logger.error(f"Error fetching admin users: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")

@app.post("/api/admin/users")
async def create_admin_user(request: Request, current_admin: dict = Depends(get_current_admin)):
    """Создание нового администратора"""
    try:
        check_super_admin(current_admin)
        data = await request.json()
        user_data = AdminUserCreate(**data)
        
        # Проверяем существование пользователя
        existing = await AdminService.get_user_by_username(user_data.username)
        if existing:
            raise HTTPException(400, "Пользователь с таким именем уже существует")
        
        user = await AdminService.create_user(user_data)
        return {"success": True, "user": user, "message": "Администратор создан"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating admin user: {e}")
        raise HTTPException(500, "Внутренняя ошибка сервера")

@app.put("/api/admin/users/{user_id}")
async def update_admin_user(user_id: int, request: Request, current_admin: dict = Depends(get_current_admin)):
    """Обновление администратора"""
    try:
        check_super_admin(current_admin)
        data = await request.json()
        user_data = AdminUserUpdate(**data)
        
        user = await AdminService.update_user(user_id, user_data)
        if not user:
            raise HTTPException(404, "Пользователь не найден")
        
        return {"success": True, "user": user, "message": "Администратор обновлен"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating admin user: {e}")
        raise HTTPException(500, "Внутренняя ошибка сервера")

@app.delete("/api/admin/users/{user_id}")
async def delete_admin_user(user_id: int, current_admin: dict = Depends(get_current_admin)):
    """Удаление администратора"""
    try:
        check_super_admin(current_admin)
        if user_id == current_admin["user_id"]:
            raise HTTPException(400, "Нельзя удалить самого себя")
        
        success = await AdminService.delete_user(user_id)
        if not success:
            raise HTTPException(404, "Пользователь не найден")
        
        return {"success": True, "message": "Администратор удален"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting admin user: {e}")
        raise HTTPException(500, "Внутренняя ошибка сервера")

# API для чата администраторов
@app.get("/api/admin/chat/messages")
async def get_chat_messages(current_admin: dict = Depends(get_current_admin)):
    """Получение сообщений чата"""
    try:
        messages = await AdminChatService.get_recent_messages(50)
        return {"messages": messages}
    except Exception as e:
        logger.error(f"Error fetching chat messages: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")

@app.post("/api/admin/chat/messages")
async def create_chat_message(request: Request, current_admin: dict = Depends(get_current_admin)):
    """Создание сообщения в чате"""
    try:
        data = await request.json()
        message_data = AdminChatMessageCreate(**data)
        
        message = await AdminChatService.create_message(
            current_admin["user_id"], 
            message_data.message
        )
        
        return {"success": True, "message": message}
        
    except Exception as e:
        logger.error(f"Error creating chat message: {e}")
        raise HTTPException(500, "Внутренняя ошибка сервера")

# API для профиля
@app.put("/api/admin/profile")
async def update_profile(request: Request, current_admin: dict = Depends(get_current_admin)):
    """Обновление профиля текущего пользователя"""
    try:
        data = await request.json()
        
        # Только супер-админ может менять свою роль
        if "role" in data and current_admin["role"] != "super_admin":
            del data["role"]
        
        user_data = AdminUserUpdate(**data)
        user = await AdminService.update_user(current_admin["user_id"], user_data)
        
        return {"success": True, "user": user, "message": "Профиль обновлен"}
        
    except Exception as e:
        logger.error(f"Error updating profile: {e}")
        raise HTTPException(500, "Внутренняя ошибка сервера")

@app.put("/api/admin/profile/password")
async def change_password(request: Request, current_admin: dict = Depends(get_current_admin)):
    """Смена пароля"""
    try:
        data = await request.json()
        current_password = data.get("current_password")
        new_password = data.get("new_password")
        
        if not current_password or not new_password:
            raise HTTPException(400, "Необходимо указать текущий и новый пароль")
        
        success = await AdminService.change_password(
            current_admin["user_id"], 
            current_password, 
            new_password
        )
        
        if not success:
            raise HTTPException(400, "Неверный текущий пароль")
        
        return {"success": True, "message": "Пароль изменен"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error changing password: {e}")
        raise HTTPException(500, "Внутренняя ошибка сервера")

@app.get("/api/stats")
async def get_stats(current_admin: dict = Depends(get_current_admin)):
    """Получение статистики для дашборда"""
    try:
        orders = await OrderService.list_recent_orders(1000)
        total_orders = len(orders)
        
        active_statuses = [s for s in STATUSES if "получен" not in s.lower() and "доставлен" not in s.lower()]
        active_orders = len([o for o in orders if o.status in active_statuses])
        
        all_participants = []
        for order in orders:
            participants = await ParticipantService.get_participants(order.order_id)
            all_participants.extend(participants)
        total_participants = len(set(p.username for p in all_participants))
        
        subscriptions = await SubscriptionService.get_all_subscriptions()
        total_subscriptions = len(subscriptions)
        
        return {
            "total_orders": total_orders,
            "active_orders": active_orders,
            "total_participants": total_participants,
            "total_subscriptions": total_subscriptions
        }
    except Exception as e:
        logger.error(f"Error fetching stats: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")

# Middleware для проверки аутентификации
@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    # Пропускаем страницу логина и статические файлы
    if request.url.path in ["/admin/login", "/admin/logout"] or request.url.path.startswith("/admin/static"):
        return await call_next(request)
    
    # Проверяем токен для защищенных страниц
    token = request.cookies.get("admin_token")
    if not token:
        return RedirectResponse(url="/admin/login")
    
    payload = verify_token(token)
    if not payload:
        response = RedirectResponse(url="/admin/login")
        response.delete_cookie("admin_token")
        return response
    
    # Добавляем информацию о пользователе в запрос
    request.state.admin_user = payload
    
    response = await call_next(request)
    return response

# Существующие API endpoints с новой аутентификацией
@app.get("/api/orders")
async def get_orders(
    status: Optional[str] = None,
    country: Optional[str] = None,
    limit: int = Query(50, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    current_admin: dict = Depends(get_current_admin)
):
    """API для получения списка заказов с пагинацией"""
    try:
        # Получаем все заказы для фильтрации
        if status:
            orders = await OrderService.list_orders_by_status([status])
        else:
            # Получаем все заказы для правильного подсчета total
            orders = await OrderService.list_recent_orders(10000)
        
        # Фильтрация по стране
        if country:
            orders = [o for o in orders if o.country == country.upper()]
        
        # Общее количество заказов (после фильтрации)
        total_orders = len(orders)
        
        # Пагинация
        paginated_orders = orders[offset:offset + limit]
        
        # Convert orders to dict for JSON serialization
        orders_data = []
        for order in paginated_orders:
            order_data = serialize_model(order)
            # Ensure datetime fields are serializable
            if order_data.get('created_at') and isinstance(order_data['created_at'], datetime):
                order_data['created_at'] = order_data['created_at'].isoformat()
            if order_data.get('updated_at') and isinstance(order_data['updated_at'], datetime):
                order_data['updated_at'] = order_data['updated_at'].isoformat()
            orders_data.append(order_data)
        
        return {
            "orders": orders_data,
            "total": total_orders,
            "has_more": total_orders > offset + limit,
            "offset": offset,
            "limit": limit
        }
    except Exception as e:
        logger.error(f"Error fetching orders: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")

@app.get("/api/orders/{order_id}")
async def get_order(order_id: str, current_admin: dict = Depends(get_current_admin)):
    """API для получения информации о заказе"""
    try:
        order = await OrderService.get_order(order_id)
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        
        participants = await ParticipantService.get_participants(order_id)
        subscriptions = await SubscriptionService.get_all_subscriptions()
        order_subs = [s for s in subscriptions if s.order_id == order_id]
        
        # Convert to dict for JSON serialization
        order_data = serialize_model(order)
        if order_data.get('created_at') and isinstance(order_data['created_at'], datetime):
            order_data['created_at'] = order_data['created_at'].isoformat()
        if order_data.get('updated_at') and isinstance(order_data['updated_at'], datetime):
            order_data['updated_at'] = order_data['updated_at'].isoformat()
        
        participants_data = []
        for participant in participants:
            participant_data = serialize_model(participant)
            if participant_data.get('created_at') and isinstance(participant_data['created_at'], datetime):
                participant_data['created_at'] = participant_data['created_at'].isoformat()
            if participant_data.get('updated_at') and isinstance(participant_data['updated_at'], datetime):
                participant_data['updated_at'] = participant_data['updated_at'].isoformat()
            participants_data.append(participant_data)
        
        return {
            "order": order_data,
            "participants": participants_data,
            "subscribers": len(order_subs)
        }
    except Exception as e:
        logger.error(f"Error fetching order {order_id}: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")

@app.post("/api/orders/create")
async def create_order_api(
    request: Request,
    current_admin: dict = Depends(get_current_admin)
):
    """Создание нового заказа"""
    try:
        data = await request.json()
        
        # Проверяем существование заказа
        existing = await OrderService.get_order(data['order_id'])
        if existing:
            raise HTTPException(400, "Заказ с таким ID уже существует")
        
        order = Order(
            order_id=data['order_id'],
            client_name=data['client_name'],
            country=data['country'].upper(),
            status=data['status'],
            note=data.get('note', '')
        )
        
        success = await OrderService.add_order(order)
        if not success:
            raise HTTPException(500, "Ошибка при создании заказа")
        
        # Добавляем участников
        from app.utils.validators import extract_usernames
        usernames = extract_usernames(data['client_name'])
        if usernames:
            await ParticipantService.ensure_participants(data['order_id'], usernames)
        
        return {"success": True, "message": "Заказ успешно создан"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating order: {e}")
        raise HTTPException(500, "Внутренняя ошибка сервера")

@app.put("/api/orders/{order_id}")
async def update_order_api(
    order_id: str,
    request: Request,
    current_admin: dict = Depends(get_current_admin)
):
    """Обновление заказа"""
    try:
        order = await OrderService.get_order(order_id)
        if not order:
            raise HTTPException(404, "Заказ не найден")
        
        data = await request.json()
        
        # Сохраняем старый статус для проверки изменений
        old_status = order.status
        
        # Обновляем поля
        update_data = {}
        if data.get('client_name') is not None:
            update_data["client_name"] = data['client_name']
        if data.get('country') is not None:
            update_data["country"] = data['country'].upper()
        if data.get('status') is not None:
            update_data["status"] = data['status']
        if data.get('note') is not None:
            update_data["note"] = data['note']
        
        if update_data:
            success = await OrderService.update_order(order_id, update_data)
            if not success:
                raise HTTPException(500, "Ошибка при обновлении заказа")
        
            # Отправляем уведомления если статус изменился
            if 'status' in update_data and update_data['status'] != old_status:
                await OrderService._send_status_notifications(order_id, update_data['status'])
        
        return {"success": True, "message": "Заказ обновлен"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating order: {e}")
        raise HTTPException(500, "Внутренняя ошибка сервера")

@app.delete("/api/orders/{order_id}")
async def delete_order_api(
    order_id: str,
    current_admin: dict = Depends(get_current_admin)
):
    """Удаление заказа"""
    try:
        # Проверяем существование заказа
        order = await OrderService.get_order(order_id)
        if not order:
            raise HTTPException(404, "Заказ не найден")
        
        success = await OrderService.delete_order(order_id)
        if not success:
            raise HTTPException(500, "Ошибка при удалении заказа")
        
        return {"success": True, "message": "Заказ удален"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting order: {e}")
        raise HTTPException(500, "Внутренняя ошибка сервера")

@app.get("/api/participants")
async def get_participants(
    order_id: Optional[str] = None,
    paid: Optional[bool] = None,
    search: Optional[str] = None,
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    current_admin: dict = Depends(get_current_admin)
):
    """API для получения списка участников с оптимизированной пагинацией"""
    try:
        # Используем новый метод для получения участников с пагинацией на уровне БД
        result = await ParticipantService.get_participants_paginated(
            order_id=order_id,
            paid=paid,
            search=search,
            limit=limit,
            offset=offset
        )
        
        # Convert to dict for JSON serialization
        participants_data = []
        for participant in result["participants"]:
            participant_data = serialize_model(participant)
            if participant_data.get('created_at') and isinstance(participant_data['created_at'], datetime):
                participant_data['created_at'] = participant_data['created_at'].isoformat()
            if participant_data.get('updated_at') and isinstance(participant_data['updated_at'], datetime):
                participant_data['updated_at'] = participant_data['updated_at'].isoformat()
            participants_data.append(participant_data)
        
        return {
            "participants": participants_data,
            "total": result["total"],
            "has_more": result["has_more"],
            "offset": offset,
            "limit": limit
        }
    except Exception as e:
        logger.error(f"Error fetching participants: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")

@app.put("/api/participants/{order_id}/{username}/paid")
async def update_participant_paid(
    order_id: str,
    username: str,
    request: Request,
    current_admin: dict = Depends(get_current_admin)
):
    """Изменение статуса оплаты участника"""
    try:
        # Используем toggle метод вместо получения данных из тела
        success = await ParticipantService.toggle_participant_paid(order_id, username)
        if not success:
            raise HTTPException(400, "Не удалось обновить статус оплаты")
        
        return {"success": True, "message": "Статус оплаты обновлен"}
        
    except Exception as e:
        logger.error(f"Error updating participant payment: {e}")
        raise HTTPException(500, "Внутренняя ошибка сервера")

@app.post("/api/broadcast/unpaid")
async def broadcast_unpaid(
    request: Request,
    current_admin: dict = Depends(get_current_admin)
):
    """Рассылка уведомлений неплательщикам"""
    try:
        # Проверяем, что тело запроса не пустое
        body = await request.body()
        if not body:
            raise HTTPException(400, "Empty request body")
        
        try:
            data = await request.json()
        except json.JSONDecodeError as e:
            logger.error(f"JSON decode error: {e}")
            raise HTTPException(400, "Invalid JSON format")
            
        message = data.get('message', '')
        
        if not message:
            raise HTTPException(400, "Сообщение не может быть пустым")
        
        # Получаем всех неплательщиков
        from app.services.order_service import ParticipantService
        unpaid_grouped = await ParticipantService.get_all_unpaid_grouped()
        
        if not unpaid_grouped:
            return {
                "success": True, 
                "message": "Нет неплательщиков для рассылки",
                "result": {
                    "sent": 0,
                    "failed": 0, 
                    "total": 0
                }
            }
        
        # Собираем все username
        all_usernames = []
        for usernames in unpaid_grouped.values():
            all_usernames.extend(usernames)
        
        # Получаем user_id по username
        user_ids = await AddressService.get_user_ids_by_usernames(all_usernames)
        
        sent_count = 0
        failed_count = 0
        
        # Отправляем сообщения через Telegram бота
        for user_id in user_ids:
            try:
                # Импортируем бота здесь чтобы избежать циклических импортов
                from app.webhook import application
                await application.bot.send_message(
                    chat_id=user_id,
                    text=message,
                    parse_mode='HTML'
                )
                sent_count += 1
            except Exception as e:
                logger.error(f"Error sending message to {user_id}: {e}")
                failed_count += 1
        
        return {
            "success": True,
            "message": "Рассылка завершена",
            "result": {
                "sent": sent_count,
                "failed": failed_count,
                "total": len(user_ids)
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error broadcasting to unpaid: {e}")
        raise HTTPException(500, "Внутренняя ошибка сервера")

@app.post("/api/broadcast/all")
async def broadcast_all(
    request: Request,
    current_admin: dict = Depends(get_current_admin)
):
    """Рассылка сообщения всем пользователям"""
    try:
        body = await request.body()
        if not body:
            raise HTTPException(400, "Empty request body")
        
        try:
            data = await request.json()
        except json.JSONDecodeError as e:
            logger.error(f"JSON decode error: {e}")
            raise HTTPException(400, "Invalid JSON format")
            
        message = data.get('message', '')
        
        if not message:
            raise HTTPException(400, "Сообщение не может быть пустым")
        
        # Получаем всех пользователей с адресами
        async with db.pool.acquire() as conn:
            rows = await conn.fetch("SELECT DISTINCT user_id FROM addresses")
            user_ids = [row['user_id'] for row in rows]
        
        sent_count = 0
        failed_count = 0
        
        # Отправляем сообщения через Telegram бота
        for user_id in user_ids:
            try:
                from app.webhook import application
                await application.bot.send_message(
                    chat_id=user_id,
                    text=message,
                    parse_mode='HTML'
                )
                sent_count += 1
            except Exception as e:
                logger.error(f"Error sending message to {user_id}: {e}")
                failed_count += 1
        
        return {
            "success": True, 
            "message": "Рассылка завершена",
            "result": {
                "sent": sent_count,
                "failed": failed_count,
                "total": len(user_ids)
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error broadcasting to all users: {e}")
        raise HTTPException(500, "Внутренняя ошибка сервера")

@app.post("/api/broadcast/reminder")
async def send_reminder(
    request: Request,
    current_admin: dict = Depends(get_current_admin)
):
    """Отправка напоминания конкретному пользователю"""
    try:
        data = await request.json()
        message = data.get('message', '')
        usernames = data.get('usernames', [])
        
        if not message or not usernames:
            raise HTTPException(400, "Сообщение и список пользователей обязательны")
        
        # Получаем user_id по username
        user_ids = await AddressService.get_user_ids_by_usernames(usernames)
        
        if not user_ids:
            return {
                "success": False,
                "message": "Пользователи не найдены"
            }
        
        sent_count = 0
        failed_count = 0
        
        # Отправляем сообщения через Telegram бота
        for user_id in user_ids:
            try:
                from app.webhook import application
                await application.bot.send_message(
                    chat_id=user_id,
                    text=message,
                    parse_mode='HTML'
                )
                sent_count += 1
            except Exception as e:
                logger.error(f"Error sending reminder to {user_id}: {e}")
                failed_count += 1
        
        return {
            "success": True,
            "message": f"Напоминания отправлены ({sent_count}/{len(user_ids)})",
            "result": {
                "sent": sent_count,
                "failed": failed_count,
                "total": len(user_ids)
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error sending reminders: {e}")
        raise HTTPException(500, "Внутренняя ошибка сервера")

@app.get("/api/statuses")
async def get_statuses(current_admin: dict = Depends(get_current_admin)):
    """API для получения списка статусов"""
    return {"statuses": STATUSES}

@app.get("/api/telegram/posts")
async def get_telegram_posts(
    limit: int = Query(5, ge=1, le=10),
    current_admin: dict = Depends(get_current_admin)
):
    """API для получения постов из Telegram канала"""
    try:
        from app.services.telegram_service import telegram_service
        posts = await telegram_service.get_channel_posts(limit)
        return {"posts": posts}
    except Exception as e:
        logger.error(f"Error fetching Telegram posts: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")

@app.get("/import")
async def import_orders_page(request: Request, current_admin: dict = Depends(get_current_admin)):
    """Страница импорта заказов"""
    return templates.TemplateResponse("import_orders.html", {
        "request": request,
        "current_admin": current_admin,
        "current_page": "orders",
        "statuses": STATUSES
    })

@app.post("/api/orders/bulk")
async def bulk_create_orders(
    request: Request,
    current_admin: dict = Depends(get_current_admin)
):
    """Массовое создание заказов"""
    try:
        data = await request.json()
        orders_data = data.get('orders', [])
        
        if not orders_data:
            raise HTTPException(400, "Нет данных для импорта")
        
        results = {
            "total": len(orders_data),
            "success": 0,
            "errors": 0,
            "duplicates": 0,
            "errorList": []
        }
        
        for order_data in orders_data:
            try:
                # Проверяем существование заказа
                existing = await OrderService.get_order(order_data['order_id'])
                if existing:
                    results["duplicates"] += 1
                    results["errorList"].append({
                        "order_id": order_data['order_id'],
                        "message": "Заказ уже существует"
                    })
                    continue
                
                # Создаем заказ
                order = Order(
                    order_id=order_data['order_id'],
                    client_name=order_data['client_name'],
                    country=order_data.get('country', 'RU').upper(),
                    status=order_data.get('status', 'В обработке'),
                    note=order_data.get('note', '')
                )
                
                success = await OrderService.add_order(order)
                if success:
                    # Добавляем участников
                    from app.utils.validators import extract_usernames
                    usernames = extract_usernames(order_data['client_name'])
                    if usernames:
                        await ParticipantService.ensure_participants(order_data['order_id'], usernames)
                    
                    # Отправляем уведомление клиенту
                    await send_order_created_notification(order, usernames)
                    
                    results["success"] += 1
                else:
                    raise Exception("Ошибка при создании заказа")
                    
            except Exception as e:
                results["errors"] += 1
                results["errorList"].append({
                    "order_id": order_data.get('order_id', 'Unknown'),
                    "message": str(e)
                })
        
        return results
        
    except Exception as e:
        logger.error(f"Error in bulk order creation: {e}")
        raise HTTPException(500, "Внутренняя ошибка сервера")

@app.post("/api/orders/parse-excel")
async def parse_excel_file(
    file: UploadFile = File(...),
    current_admin: dict = Depends(get_current_admin)
):
    """Парсинг Excel файла с заказами"""
    try:
        # Проверяем тип файла
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(400, "Поддерживаются только файлы Excel (.xlsx, .xls)")
        
        # Читаем файл
        contents = await file.read()
        
        # Простой парсинг Excel (в реальности нужно использовать pandas или openpyxl)
        # Сейчас возвращаем заглушки для демонстрации
        orders = []
        
        # Пример данных из Excel
        sample_data = [
            {"order_id": "ORD-001", "client_name": "@user1", "country": "RU", "status": "В обработке"},
            {"order_id": "ORD-002", "client_name": "@user2", "country": "KZ", "status": "В пути"},
            {"order_id": "ORD-003", "client_name": "@user3", "country": "UZ", "status": "Доставлен"},
        ]
        
        # В реальности здесь будет парсинг Excel файла
        # orders = parse_excel_contents(contents)
        
        return {
            "orders": sample_data,
            "total": len(sample_data),
            "message": "Файл успешно распарсен (демо-данные)"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error parsing Excel file: {e}")
        raise HTTPException(500, "Ошибка при обработке файла")

async def send_order_created_notification(order: Order, usernames: List[str]):
    """Отправка уведомления клиенту при создании заказа"""
    try:
        if not usernames:
            return
        
        # Получаем user_id по username
        from app.services.user_service import AddressService
        user_ids = await AddressService.get_user_ids_by_usernames(usernames)
        
        if not user_ids:
            return
        
        # Формируем сообщение
        message = f"🎉 <b>Новый заказ создан!</b>\n\n"
        message += f"📦 <b>Заказ:</b> {order.order_id}\n"
        message += f"👤 <b>Клиент:</b> {order.client_name}\n"
        message += f"🌍 <b>Страна:</b> {order.country}\n"
        message += f"🔄 <b>Статус:</b> {order.status}\n"
        
        if order.note:
            message += f"📝 <b>Примечание:</b> {order.note}\n"
        
        message += f"\n💡 <i>Следите за обновлениями статуса!</i>"
        
        # Отправляем уведомления
        from app.webhook import application
        for user_id in user_ids:
            try:
                await application.bot.send_message(
                    chat_id=user_id,
                    text=message,
                    parse_mode='HTML'
                )
                logger.info(f"Sent order creation notification to {user_id} for order {order.order_id}")
            except Exception as e:
                logger.error(f"Error sending notification to {user_id}: {e}")
                
    except Exception as e:
        logger.error(f"Error sending order creation notifications: {e}")

# Добавим уведомление и в обычное создание заказа
@app.post("/api/orders/create")
async def create_order_api(
    request: Request,
    current_admin: dict = Depends(get_current_admin)
):
    """Создание нового заказа"""
    try:
        data = await request.json()
        
        # Проверяем существование заказа
        existing = await OrderService.get_order(data['order_id'])
        if existing:
            raise HTTPException(400, "Заказ с таким ID уже существует")
        
        order = Order(
            order_id=data['order_id'],
            client_name=data['client_name'],
            country=data['country'].upper(),
            status=data['status'],
            note=data.get('note', '')
        )
        
        success = await OrderService.add_order(order)
        if not success:
            raise HTTPException(500, "Ошибка при создании заказа")
        
        # Добавляем участников
        from app.utils.validators import extract_usernames
        usernames = extract_usernames(data['client_name'])
        if usernames:
            await ParticipantService.ensure_participants(data['order_id'], usernames)
        
        # Отправляем уведомление клиенту
        await send_order_created_notification(order, usernames)
        
        return {"success": True, "message": "Заказ успешно создан"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating order: {e}")
        raise HTTPException(500, "Внутренняя ошибка сервера")

@app.get("/orders/{order_id}", response_class=HTMLResponse)
async def view_order_page(request: Request, order_id: str, current_admin: dict = Depends(get_current_admin)):
    """Страница просмотра заказа"""
    try:
        order = await OrderService.get_order(order_id)
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        
        participants = await ParticipantService.get_participants(order_id)
        
        return templates.TemplateResponse("order_view.html", {
            "request": request,
            "current_admin": current_admin,
            "current_page": "orders",
            "order": order,
            "participants": participants
        })
    except Exception as e:
        logger.error(f"Error loading order page {order_id}: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")

@app.get("/api/reports/analytics")
async def get_analytics(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_admin: dict = Depends(get_current_admin)
):
    """API для получения аналитики"""
    try:
        # Получаем все заказы для анализа
        orders = await OrderService.list_recent_orders(1000)
        
        # Статистика по статусам
        status_stats = {}
        for status in STATUSES:
            count = len([o for o in orders if o.status == status])
            if count > 0:
                status_stats[status] = count
        
        # Статистика по странам
        country_stats = {}
        for order in orders:
            country = order.country
            country_stats[country] = country_stats.get(country, 0) + 1
        
        # Статистика по платежам
        total_participants = 0
        paid_participants = 0
        
        # Получаем всех участников
        all_participants = []
        for order in orders:
            participants = await ParticipantService.get_participants(order.order_id)
            all_participants.extend(participants)
            total_participants += len(participants)
            paid_participants += len([p for p in participants if p.paid])
        
        # Уникальные участники
        unique_participants = len(set(p.username for p in all_participants))
        
        return {
            "status_stats": status_stats,
            "country_stats": country_stats,
            "payment_stats": {
                "total": total_participants,
                "paid": paid_participants,
                "unpaid": total_participants - paid_participants
            },
            "total_orders": len(orders),
            "unique_participants": unique_participants,
            "completed_orders": len([o for o in orders if "доставлен" in o.status.lower() or "получен" in o.status.lower()])
        }
    except Exception as e:
        logger.error(f"Error generating analytics: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")

@app.get("/api/reports/export/participants")
async def export_participants(
    format: str = Query("csv", regex="^(csv|json|xlsx)$"),
    current_admin: dict = Depends(get_current_admin)
):
    """Экспорт участников"""
    try:
        # Получаем всех участников
        all_participants = []
        orders = await OrderService.list_recent_orders(1000)
        for order in orders:
            participants = await ParticipantService.get_participants(order.order_id)
            for p in participants:
                participant_data = serialize_model(p)
                participant_data["order_client_name"] = order.client_name
                participant_data["order_status"] = order.status
                participant_data["order_country"] = order.country
                all_participants.append(participant_data)
        
        if format == "xlsx":
            # Экспорт в XLSX
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Участники заказов"
            
            # Заголовки
            headers = ["Username", "ID заказа", "Имя клиента", "Статус", "Страна", "Оплачено", "Обновлен"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Данные
            for row, p in enumerate(all_participants, 2):
                ws.cell(row=row, column=1, value=f"@{p['username']}")
                ws.cell(row=row, column=2, value=p['order_id'])
                ws.cell(row=row, column=3, value=p['order_client_name'])
                ws.cell(row=row, column=4, value=p['order_status'])
                ws.cell(row=row, column=5, value=p['order_country'])
                ws.cell(row=row, column=6, value="Да" if p['paid'] else "Нет")
                ws.cell(row=row, column=7, value=p.get('updated_at', ''))
            
            # Автоматическая ширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            filename = f"participants_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return Response(
                content=buffer.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        
        elif format == "csv":
            import csv
            from io import StringIO
            
            output = StringIO()
            writer = csv.writer(output)
            
            # Заголовки
            writer.writerow(["Username", "Order ID", "Client Name", "Status", "Country", "Paid", "Updated At"])
            
            for p in all_participants:
                writer.writerow([
                    f"@{p['username']}",
                    p['order_id'],
                    p['order_client_name'],
                    p['order_status'],
                    p['order_country'],
                    "Да" if p['paid'] else "Нет",
                    p.get('updated_at', '')
                ])
            
            content = output.getvalue()
            return JSONResponse({
                "content": content,
                "filename": f"participants_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            })
            
        else:  # json
            return {
                "participants": all_participants,
                "filename": f"participants_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            }
            
    except Exception as e:
        logger.error(f"Error exporting participants: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")

@app.get("/api/reports/export/orders")
async def export_orders(
    format: str = Query("csv", regex="^(csv|json|xlsx)$"),
    current_admin: dict = Depends(get_current_admin)
):
    """Экспорт заказов"""
    try:
        orders = await OrderService.list_recent_orders(1000)
        
        if format == "xlsx":
            # Экспорт в XLSX
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Заказы"
            
            # Заголовки
            headers = ["ID заказа", "Имя клиента", "Страна", "Статус", "Примечание", "Создан", "Обновлен"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Данные
            for row, order in enumerate(orders, 2):
                ws.cell(row=row, column=1, value=order.order_id)
                ws.cell(row=row, column=2, value=order.client_name)
                ws.cell(row=row, column=3, value=order.country)
                ws.cell(row=row, column=4, value=order.status)
                ws.cell(row=row, column=5, value=order.note or "")
                ws.cell(row=row, column=6, value=order.created_at.isoformat() if order.created_at else "")
                ws.cell(row=row, column=7, value=order.updated_at.isoformat() if order.updated_at else "")
            
            # Автоматическая ширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            filename = f"orders_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return Response(
                content=buffer.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        
        elif format == "csv":
            import csv
            from io import StringIO
            
            output = StringIO()
            writer = csv.writer(output)
            
            # Заголовки
            writer.writerow(["Order ID", "Client Name", "Country", "Status", "Note", "Created At", "Updated At"])
            
            for order in orders:
                writer.writerow([
                    order.order_id,
                    order.client_name,
                    order.country,
                    order.status,
                    order.note or "",
                    order.created_at.isoformat() if order.created_at else "",
                    order.updated_at.isoformat() if order.updated_at else ""
                ])
            
            content = output.getvalue()
            return JSONResponse({
                "content": content,
                "filename": f"orders_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            })
            
        else:  # json
            orders_data = []
            for order in orders:
                order_data = serialize_model(order)
                if order_data.get('created_at') and isinstance(order_data['created_at'], datetime):
                    order_data['created_at'] = order_data['created_at'].isoformat()
                if order_data.get('updated_at') and isinstance(order_data['updated_at'], datetime):
                    order_data['updated_at'] = order_data['updated_at'].isoformat()
                orders_data.append(order_data)
            
            return {
                "orders": orders_data,
                "filename": f"orders_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            }
            
    except Exception as e:
        logger.error(f"Error exporting orders: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")


@app.get("/addresses", response_class=HTMLResponse)
async def addresses_page(request: Request, current_admin: dict = Depends(get_current_admin)):
    """Страница адресов клиентов"""
    return templates.TemplateResponse("addresses.html", {
        "request": request,
        "current_admin": current_admin,
        "current_page": "addresses"
    })


@app.get("/api/addresses")
async def get_addresses(
    limit: int = Query(50, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    current_admin: dict = Depends(get_current_admin)
):
    """API для получения списка адресов с пагинацией"""
    try:
        addresses = await AddressService.get_all_addresses()
        
        # Общее количество адресов
        total_addresses = len(addresses)
        
        # Пагинация
        paginated_addresses = addresses[offset:offset + limit]
        
        addresses_data = []
        for address in paginated_addresses:
            address_data = serialize_model(address)
            if address_data.get('created_at') and isinstance(address_data['created_at'], datetime):
                address_data['created_at'] = address_data['created_at'].isoformat()
            if address_data.get('updated_at') and isinstance(address_data['updated_at'], datetime):
                address_data['updated_at'] = address_data['updated_at'].isoformat()
            addresses_data.append(address_data)
        
        return {
            "addresses": addresses_data,
            "total": total_addresses,
            "has_more": total_addresses > offset + limit,
            "offset": offset,
            "limit": limit
        }
    except Exception as e:
        logger.error(f"Error fetching addresses: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")


@app.get("/api/addresses/export/xlsx")
async def export_addresses_xlsx(current_admin: dict = Depends(get_current_admin)):
    """Экспорт адресов в XLSX"""
    try:
        addresses = await AddressService.get_all_addresses()
        
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Адреса клиентов"
        
        # Заголовки
        headers = ["Telegram ID", "Username", "Имя", "Телефон", "Город", "Адрес", "Индекс", "Обновлен"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Данные
        for row, address in enumerate(addresses, 2):
            ws.cell(row=row, column=1, value=address.user_id)
            ws.cell(row=row, column=2, value=f"@{address.username}")
            ws.cell(row=row, column=3, value=address.full_name or "")
            ws.cell(row=row, column=4, value=address.phone or "")
            ws.cell(row=row, column=5, value=address.city or "")
            ws.cell(row=row, column=6, value=address.address or "")
            ws.cell(row=row, column=7, value=address.postcode or "")
            ws.cell(row=row, column=8, value=address.updated_at.isoformat() if address.updated_at else "")
        
        # Автоматическая ширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"addresses_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return Response(
            content=buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Error exporting addresses: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")

@app.post("/api/orders/bulk-update-status")
async def bulk_update_order_status(
    request: Request,
    current_admin: dict = Depends(get_current_admin)
):
    """Массовое обновление статусов заказов"""
    try:
        data = await request.json()
        order_ids = data.get('order_ids', [])
        new_status = data.get('status', '')
        
        if not order_ids:
            raise HTTPException(400, "Не выбраны заказы для обновления")
        
        if not new_status:
            raise HTTPException(400, "Не указан новый статус")
        
        if new_status not in STATUSES:
            raise HTTPException(400, f"Неверный статус. Допустимые значения: {STATUSES}")
        
        success_count = 0
        failed_count = 0
        
        for order_id in order_ids:
            try:
                # Получаем старый статус для проверки изменений
                old_order = await OrderService.get_order(order_id)
                old_status = old_order.status if old_order else ""
                
                success = await OrderService.update_order_status(order_id, new_status)
                if success:
                    success_count += 1
                    # Отправляем уведомления если статус изменился
                    if old_status != new_status:
                        await OrderService._send_status_notifications(order_id, new_status)
                else:
                    failed_count += 1
            except Exception as e:
                logger.error(f"Error updating order {order_id}: {e}")
                failed_count += 1
        
        message = f"Обновлено {success_count} заказов"
        if failed_count > 0:
            message += f", не удалось обновить {failed_count}"
        
        return {
            "success": True,
            "message": message,
            "updated": success_count,
            "failed": failed_count
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in bulk update order status: {e}")
        raise HTTPException(500, "Внутренняя ошибка сервера")

@app.post("/api/orders/bulk-delete")
async def bulk_delete_orders(
    request: Request,
    current_admin: dict = Depends(get_current_admin)
):
    """Массовое удаление заказов"""
    try:
        data = await request.json()
        order_ids = data.get('order_ids', [])
        
        if not order_ids:
            raise HTTPException(400, "Не выбраны заказы для удаления")
        
        success_count = 0
        failed_count = 0
        
        for order_id in order_ids:
            try:
                success = await OrderService.delete_order(order_id)
                if success:
                    success_count += 1
                else:
                    failed_count += 1
            except Exception as e:
                logger.error(f"Error deleting order {order_id}: {e}")
                failed_count += 1
        
        message = f"Удалено {success_count} заказов"
        if failed_count > 0:
            message += f", не удалось удалить {failed_count}"
        
        return {
            "success": True,
            "message": message,
            "deleted": success_count,
            "failed": failed_count
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in bulk delete orders: {e}")
        raise HTTPException(500, "Внутренняя ошибка сервера")

# Создаем директорию для аватарок
AVATAR_DIR = os.path.join(STATIC_DIR, "avatars")
os.makedirs(AVATAR_DIR, exist_ok=True)

@app.post("/api/admin/profile/avatar")
async def upload_avatar(
    avatar: UploadFile = File(...),
    current_admin: dict = Depends(get_current_admin)
):
    """Загрузка аватарки пользователя"""
    try:
        # Проверяем тип файла
        if not avatar.content_type or not avatar.content_type.startswith("image/"):
            raise HTTPException(400, "Файл должен быть изображением")
        
        # Читаем содержимое файла
        contents = await avatar.read()
        
        # Проверяем размер файла (не более 5MB)
        max_size = 5 * 1024 * 1024  # 5 MB
        if len(contents) > max_size:
            raise HTTPException(400, "Размер файла не должен превышать 5MB")
        
        # Открываем изображение через PIL
        image = Image.open(io.BytesIO(contents))
        
        # Конвертируем в RGB при необходимости
        if image.mode in ("RGBA", "LA", "P"):
            image = image.convert("RGB")
        
        # Ресайзим до 200x200
        try:
            resampling = Image.Resampling.LANCZOS
        except AttributeError:
            resampling = Image.LANCZOS
        image.thumbnail((200, 200), resampling)
        
        # Определяем расширение файла
        original_name = avatar.filename or "avatar"
        if "." in original_name:
            ext = original_name.rsplit(".", 1)[1].lower()
        else:
            ext = "jpg"
        
        if ext not in ["jpg", "jpeg", "png"]:
            ext = "jpg"
        
        # Генерируем уникальное имя файла
        filename = f"{current_admin['user_id']}_{uuid.uuid4().hex[:8]}.{ext}"
        save_path = os.path.join(AVATAR_DIR, filename)
        
        # Пересоздаём директорию на всякий случай (если вдруг её удалили)
        os.makedirs(AVATAR_DIR, exist_ok=True)
        
        # Сохраняем изображение
        image.save(save_path, "JPEG" if ext in ["jpg", "jpeg"] else "PNG", quality=85)
        
        # Обновляем аватарку в базе данных
        avatar_url = f"/static/avatars/{filename}"
        user = await AdminService.update_user(
            current_admin["user_id"],
            AdminUserUpdate(avatar_url=avatar_url)
        )
        
        return {
            "success": True,
            "message": "Аватарка успешно обновлена",
            "avatar_url": avatar_url
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error uploading avatar: {e}")
        raise HTTPException(500, "Ошибка при загрузке аватарки")
